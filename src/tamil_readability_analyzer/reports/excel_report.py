"""Excel report rendering."""

import io


def generate_readability_excel(row, results, distribution, proper_nouns, sent_data, meaning=None):
    """Build the readability workbook and return XLSX bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    book_name = row['book_name']
    tss = sent_data.get('target', {})

    wb = Workbook()
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1D4E89')
    gf = PatternFill('solid', fgColor='D4EDDA')
    af = PatternFill('solid', fgColor='FFF3CD')
    rf = PatternFill('solid', fgColor='F8D7DA')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def hdr_row(ws, headers, row_num=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = thin

    ws1 = wb.active
    ws1.title = 'Readability Summary'
    ws1['A1'] = f'Tamil Readability Report - {book_name}'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = (
        f'Analyzed: {row["analyzed_at"][:19]}  |  '
        f'Total words: {row["total_words"]:,}  |  '
        f'Unique stems: {row.get("unique_stems","-")}  |  '
        f'Proper nouns: {len(proper_nouns)}  |  '
        f'Target avg sentence: {tss.get("avg","-")} words  |  '
        f'Target max sentence: {tss.get("max","-")} words'
    )
    ws1['A2'].font = Font(italic=True, color='555555', size=9)
    ws1.append([])
    headers1 = [
        'Standard (cumulative)',
        'Total unique words in book',
        'Known words (Std 1-N)',
        '% known',
        'New words (new to student)',
        '% new',
        'Verdict',
        'Grade max sentence',
        'Book sentences over grade max',
    ]
    hdr_row(ws1, headers1, 4)
    for r in results:
        pct = r['known_pct']
        verdict = 'Easy' if pct >= 90 else 'Readable' if pct >= 80 else 'Challenging' if pct >= 60 else 'Very Hard'
        ws1.append([
            f"Std 1-{r['grade']}",
            r['total_unique_book_words'],
            r['known_words'],
            f"{r['known_pct']}%",
            r['new_words'],
            f"{r['new_pct']}%",
            verdict,
            r.get('grade_sent_max', '-'),
            r.get('target_sentences_over_max', '-'),
        ])
        rn = ws1.max_row
        fill = gf if pct >= 80 else af if pct >= 60 else rf
        for col in range(1, 10):
            c = ws1.cell(row=rn, column=col)
            c.fill = fill
            c.border = thin
            c.alignment = Alignment(horizontal='center')
    for col, width in enumerate([16, 22, 20, 10, 22, 10, 14, 16, 22], 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    ws_dist = wb.create_sheet('Word Distribution')
    hdr_row(ws_dist, ['Class', 'Word count', '% of book vocabulary', 'Word'])
    for drow in distribution:
        label = drow.get('label') or f"Std {drow.get('grade')}"
        words = drow.get('words') or []
        if words:
            for i, word in enumerate(words):
                ws_dist.append([
                    label if i == 0 else '',
                    drow.get('word_count', 0) if i == 0 else '',
                    f"{drow.get('word_pct', 0)}%" if i == 0 else '',
                    word,
                ])
        else:
            ws_dist.append([label, drow.get('word_count', 0), f"{drow.get('word_pct', 0)}%", ''])
    ws_dist.column_dimensions['A'].width = 18
    ws_dist.column_dimensions['B'].width = 14
    ws_dist.column_dimensions['C'].width = 20
    ws_dist.column_dimensions['D'].width = 28

    ws2 = wb.create_sheet('New Words (new to student)')
    hdr_row(ws2, ['Standard (Std 1-N)', 'New word count', '% new', 'Word'])
    for r in results:
        for i, word in enumerate(r.get('new_word_list', [])):
            ws2.append([
                f"Std 1-{r['grade']}" if i == 0 else '',
                r['new_words'] if i == 0 else '',
                f"{r['new_pct']}%" if i == 0 else '',
                word,
            ])
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 28

    ws3 = wb.create_sheet('New Words Detail')
    hdr_row(ws3, ['Standard (Std 1-N)', 'New words (new to student)', 'Word'])
    for r in results:
        for i, word in enumerate(r.get('new_word_list', [])):
            ws3.append([
                f"Std 1-{r['grade']}" if i == 0 else '',
                r['new_words'] if i == 0 else '',
                word,
            ])
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 28

    ws4 = wb.create_sheet('Sentence Analysis')
    ws4['A1'] = (
        f'Target book: avg={tss.get("avg","-")} words/sent | '
        f'max={tss.get("max","-")} | median={tss.get("median","-")} | '
        f'sentences={tss.get("total_sentences",0):,}'
    )
    ws4['A1'].font = Font(bold=True)
    ws4.append([])
    hdr_row(ws4, [
        'Standard', 'Grade Max Sentence', 'Grade Avg Sentence',
        'Target Avg Sentence', 'Sentences Over Grade Max', '% Over Max',
    ], 3)
    for r in results:
        ws4.append([
            f"Std {r['grade']}",
            r.get('grade_sent_max', '-'),
            r.get('grade_sent_avg', '-'),
            tss.get('avg', '-'),
            r.get('target_sentences_over_max', '-'),
            f"{r.get('target_pct_over_max','-')}%",
        ])
    for col, width in enumerate([12, 18, 18, 18, 22, 12], 1):
        ws4.column_dimensions[get_column_letter(col)].width = width

    if meaning and meaning.get('enabled'):
        wsm = wb.create_sheet('Meaning Appropriateness')
        wsm['A1'] = (
            f"Meaning-level target: Std {meaning.get('target_grade','-')} | "
            f"Score: {meaning.get('appropriateness_pct','-')}% | "
            f"Flags: {meaning.get('flagged_count',0)}"
        )
        wsm['A1'].font = Font(bold=True)
        wsm.append([])
        hdr_row(wsm, ['Item', 'Type', 'Frequency', 'Detected level', 'Gap', 'Severity', 'Concept'], 3)
        for f in meaning.get('flagged', []):
            wsm.append([
                f.get('item'), f.get('type'), f.get('freq'), f.get('level'),
                f.get('gap'), f.get('severity'), f.get('concept'),
            ])
        for col, width in enumerate([30, 12, 12, 15, 8, 18, 22], 1):
            wsm.column_dimensions[get_column_letter(col)].width = width

    ws5 = wb.create_sheet('Proper Nouns')
    ws5['A1'] = 'Proper nouns counted as known at all grade levels'
    ws5['A1'].font = Font(bold=True)
    ws5.append(['Word'])
    ws5['A2'].font = Font(bold=True)
    for word in proper_nouns:
        ws5.append([word])
    ws5.column_dimensions['A'].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
