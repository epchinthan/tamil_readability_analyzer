"""
word_library.py — Age-wise Tamil Word Library

A standalone vocabulary reference database fed by multiple sources:

  1. School textbooks        — exact grade from loaded DB (most reliable)
  2. Children's books/PDFs  — grade inferred from readability analysis
  3. Tamil Wikipedia         — grade inferred from co-occurrence (one-time download)
  4. Tamil Wiktionary        — definitions, parts-of-speech (one-time download)
  5. Manual teacher entry    — authoritative override, stored permanently

Each word has: grade_level, frequency, concept_category, example_sentence,
part_of_speech, definition, source list, confirmed flag.

The library is stored in a separate SQLite DB (word_library.db) so it
never interferes with the main analyzer DB.
"""
from __future__ import annotations

import json, os, re, sqlite3, datetime, bz2, gzip, html
import urllib.request
from collections import defaultdict, Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

TAMIL_RE   = re.compile(r'[\u0B80-\u0BFF]{2,}')
SENT_RE    = re.compile(r'(?<=[.!?।\u0964\u0965])\s+|\n+')
LIB_DB     = 'word_library.db'
WIKI_DB    = 'data/wiki_corpus.db'

# Concept seed words — used for category classification
CONCEPT_SEEDS: Dict[str, List[str]] = {
    'language_literature': [
        'தமிழ்','மொழி','இலக்கியம்','இலக்கணம்','கவிதை','பாடல்','சொல்','எழுத்து','பேச்சு',
        'உரை','கதை','நூல்','புத்தகம்','வரி','பா','பாட்டு','கவி','செய்யுள்',
    ],
    'nature_environment': [
        'மரம்','காடு','நீர்','காற்று','மழை','சுற்றுச்சூழல்','மாசு','விலங்கு','பறவை',
        'நிலம்','பூமி','வானம்','கடல்','ஆறு','மலை','பூ','இலை','தாவரம்','பயிர்','வயல்',
        'சூரியன்','நிலவு','நட்சத்திரம்','மேகம்','மண்','புல்','மீன்','நாய்','பூனை',
    ],
    'science_technology': [
        'விஞ்ஞானம்','அறிவியல்','தொழில்நுட்பம்','ஆராய்ச்சி','கண்டுபிடிப்பு','வேதியியல்',
        'இயற்பியல்','உயிரியல்','மின்சாரம்','கணினி','இணையம்','ஆற்றல்','வெப்பம்',
        'ஒளி','ஒலி','அழுத்தம்','திசைவேகம்','உராய்வு','அணு','மூலக்கூறு',
    ],
    'society_civics': [
        'நாடு','அரசு','சமூகம்','உரிமை','சட்டம்','தேர்தல்','குடிமை','சமத்துவம்',
        'தேசியம்','குடும்பம்','கல்வி','பொருளாதாரம்','வேலை','தொழில்','வணிகம்',
        'வரி','நீதி','அமைதி','போர்','ஊர்','நகரம்',
    ],
    'health_body': [
        'உடல்','நலம்','உணவு','நோய்','சுகாதாரம்','பாதுகாப்பு','மருத்துவம்','உறுப்பு',
        'இரத்தம்','எலும்பு','தசை','மூச்சு','உள்ளம்','மனம்','தூக்கம்','உடற்பயிற்சி',
        'காய்கறி','பழம்','தண்ணீர்','உணர்வு',
    ],
    'math_logic': [
        'எண்','கூட்டல்','கழித்தல்','வடிவம்','அளவு','கோணம்','வட்டம்','பெருக்கல்',
        'வகுத்தல்','சதவீதம்','பின்னம்','கணக்கு','நீளம்','அகலம்','உயரம்','எடை',
        'நேரம்','காலம்','தூரம்','முக்கோணம்','சதுரம்','கன்',
    ],
    'history_culture': [
        'வரலாறு','பண்பாடு','மரபு','சங்க','வள்ளல்','புலவர்','கலை','திருவிழா',
        'கோட்டை','அரண்மனை','மன்னன்','இராச்சியம்','போர்','வீரம்','தியாகம்',
        'தொல்லியல்','பழமை','கல்வெட்டு','சிலை','கோவில்','மசூதி','தேவாலயம்',
    ],
    'daily_life': [
        'வீடு','சோறு','தண்ணீர்','தூக்கம்','விளையாட்டு','நண்பன்','அம்மா','அப்பா',
        'தாத்தா','பாட்டி','அக்கா','அண்ணன்','தங்கை','தம்பி','பள்ளி','ஆசிரியர்',
        'சாலை','கடை','பணம்','உடை','படம்','பயணம்',
    ],
}

STOPWORDS = {
    'ஒரு','இந்த','அந்த','இது','அது','எது','என்','உன்','தன்','நம்','நான்','நீ',
    'அவன்','அவள்','அவர்','அவர்கள்','அவை','இவை','இவர்','இவர்கள்',
    'மற்றும்','ஆகிய','என','என்று','என்ற','உள்ள','இல்லை','ஆம்','ஆகும்',
    'வேண்டும்','முதல்','வரை','அல்லது','ஆனால்','பின்','முன்','மேல்','கீழ்',
    'உடன்','வழி','போல்','போன்ற','மிக','பல','எல்லாம்','யார்','என்ன',
    'கொண்டு','செய்து','வரும்','இருக்கும்','உண்டு',
}

CORE_WORD_GRADES = {
    word: 1 for word in STOPWORDS
}
CORE_WORD_GRADES.update({
    'நான்': 1, 'நீ': 1, 'நாம்': 1, 'நாங்கள்': 1, 'என்': 1, 'எங்கள்': 1,
    'அம்மா': 1, 'அப்பா': 1, 'வீடு': 1, 'பள்ளி': 1, 'புத்தகம்': 1,
})

GRADE_LABELS = {
    1:'Standard 1', 2:'Standard 2', 3:'Standard 3', 4:'Standard 4',
    5:'Standard 5', 6:'Standard 6', 7:'Standard 7', 8:'Standard 8',
    9:'Standard 9', 10:'Standard 10', 11:'Standard 11', 12:'Standard 12',
}

# ── DB setup ──────────────────────────────────────────────────────────────────

def get_lib_db() -> sqlite3.Connection:
    conn = sqlite3.connect(LIB_DB, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute('PRAGMA cache_size=-32000')
    return conn


def init_library_db() -> None:
    conn = get_lib_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS word_library (
            stem            TEXT PRIMARY KEY,
            display_word    TEXT,
            grade_level     INTEGER,
            grade_source    TEXT DEFAULT 'inferred',
            frequency       INTEGER DEFAULT 0,
            concept         TEXT DEFAULT 'general',
            example         TEXT,
            example_source  TEXT,
            part_of_speech  TEXT,
            definition      TEXT,
            confirmed       INTEGER DEFAULT 0,
            sources_json    TEXT DEFAULT '[]',
            added_at        TEXT,
            updated_at      TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_lib_grade   ON word_library(grade_level);
        CREATE INDEX IF NOT EXISTS idx_lib_concept ON word_library(concept);
        CREATE INDEX IF NOT EXISTS idx_lib_confirmed ON word_library(confirmed);
        CREATE INDEX IF NOT EXISTS idx_lib_display ON word_library(display_word);

        CREATE TABLE IF NOT EXISTS library_sources (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT UNIQUE,
            source_type TEXT,
            grade       INTEGER,
            file_path   TEXT,
            word_count  INTEGER DEFAULT 0,
            added_at    TEXT,
            status      TEXT DEFAULT 'active'
        );

        CREATE TABLE IF NOT EXISTS library_import_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            source_name TEXT,
            words_added INTEGER,
            words_updated INTEGER,
            started_at  TEXT,
            finished_at TEXT,
            status      TEXT
        );
    ''')
    conn.commit()
    conn.close()


def get_wiki_db() -> sqlite3.Connection:
    Path(WIKI_DB).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(WIKI_DB, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    return conn


def init_wiki_db() -> None:
    conn = get_wiki_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS wiki_words (
            stem            TEXT PRIMARY KEY,
            display_word    TEXT,
            inferred_grade  INTEGER,
            confidence      REAL DEFAULT 0,
            frequency       INTEGER DEFAULT 0,
            example         TEXT,
            article_count   INTEGER DEFAULT 0,
            source_dump     TEXT,
            grade_reason    TEXT,
            grade_estimated_at TEXT,
            updated_at      TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_wiki_grade ON wiki_words(inferred_grade);
        CREATE INDEX IF NOT EXISTS idx_wiki_freq ON wiki_words(frequency);

        CREATE TABLE IF NOT EXISTS wiki_imports (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            dump_path      TEXT,
            dump_type      TEXT,
            article_count  INTEGER DEFAULT 0,
            stem_count     INTEGER DEFAULT 0,
            imported_at    TEXT,
            status         TEXT
        );
    ''')
    for col, spec in [
        ('grade_reason', 'TEXT'),
        ('grade_estimated_at', 'TEXT'),
    ]:
        try:
            conn.execute(f'ALTER TABLE wiki_words ADD COLUMN {col} {spec}')
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()


def get_core_word_info(stem: str, display_word: str = '') -> Optional[Dict]:
    """Return built-in core/function-word metadata for words not stored as glossary entries."""
    candidates = [stem, display_word]
    for word in candidates:
        if word in CORE_WORD_GRADES:
            return {
                'stem': stem or word,
                'display_word': display_word or word,
                'grade_level': CORE_WORD_GRADES[word],
                'grade_source': 'core_tamil',
                'frequency': 0,
                'concept': 'daily_life',
                'confirmed': 1,
                'definition': 'Core Tamil function/basic word',
            }
    return None


def upsert_wiki_word(
    conn: sqlite3.Connection,
    *,
    stem: str,
    display_word: str,
    inferred_grade: int,
    confidence: float,
    frequency: int,
    example: Optional[str],
    source_dump: str,
) -> None:
    now = datetime.datetime.now().isoformat()
    existing = conn.execute('SELECT * FROM wiki_words WHERE stem=?', (stem,)).fetchone()
    if existing is None:
        conn.execute('''
            INSERT INTO wiki_words
              (stem, display_word, inferred_grade, confidence, frequency,
               example, article_count, source_dump, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (stem, display_word, inferred_grade, confidence, frequency,
              example, 1, source_dump, now))
        return
    existing = dict(existing)
    old_freq = int(existing.get('frequency') or 0)
    new_freq = old_freq + frequency
    old_conf = float(existing.get('confidence') or 0)
    keep_new_grade = confidence > old_conf or (
        confidence == old_conf and inferred_grade < int(existing.get('inferred_grade') or 99)
    )
    conn.execute('''
        UPDATE wiki_words SET
          display_word=?,
          inferred_grade=?,
          confidence=?,
          frequency=?,
          example=?,
          article_count=?,
          source_dump=?,
          updated_at=?
        WHERE stem=?
    ''', (
        display_word or existing.get('display_word') or stem,
        inferred_grade if keep_new_grade else existing.get('inferred_grade'),
        confidence if keep_new_grade else old_conf,
        new_freq,
        existing.get('example') or example,
        int(existing.get('article_count') or 0) + 1,
        source_dump,
        now,
        stem,
    ))


def get_wiki_stats() -> Dict:
    init_wiki_db()
    conn = get_wiki_db()
    total = conn.execute('SELECT COUNT(*) FROM wiki_words').fetchone()[0]
    tokens = conn.execute('SELECT COALESCE(SUM(frequency), 0) FROM wiki_words').fetchone()[0]
    by_grade = conn.execute('''
        SELECT inferred_grade AS grade, COUNT(*) AS cnt
        FROM wiki_words GROUP BY inferred_grade ORDER BY inferred_grade
    ''').fetchall()
    by_reason = conn.execute('''
        SELECT COALESCE(grade_reason, 'not_estimated') AS reason, COUNT(*) AS cnt
        FROM wiki_words GROUP BY COALESCE(grade_reason, 'not_estimated')
        ORDER BY cnt DESC LIMIT 12
    ''').fetchall()
    estimated = conn.execute('''
        SELECT COUNT(*) FROM wiki_words
        WHERE grade_estimated_at IS NOT NULL
    ''').fetchone()[0]
    avg_conf = conn.execute('''
        SELECT COALESCE(AVG(confidence), 0) FROM wiki_words
        WHERE grade_estimated_at IS NOT NULL
    ''').fetchone()[0]
    imports = conn.execute('SELECT * FROM wiki_imports ORDER BY imported_at DESC LIMIT 10').fetchall()
    top = conn.execute('''
        SELECT stem, display_word, inferred_grade, confidence, grade_reason, frequency
        FROM wiki_words ORDER BY frequency DESC LIMIT 20
    ''').fetchall()
    conn.close()
    return {
        'database': WIKI_DB,
        'total_stems': total,
        'total_tokens': tokens,
        'estimated_count': estimated,
        'avg_confidence': round(float(avg_conf or 0), 2),
        'by_grade': [dict(r) for r in by_grade],
        'by_reason': [dict(r) for r in by_reason],
        'imports': [dict(r) for r in imports],
        'top_words': [dict(r) for r in top],
    }


FORMAL_GRADE_SUFFIXES = ('த்துவம்', 'வியல்', 'வாதம்', 'வாதி', 'முறை', 'நிலை', 'பாடு')
TECHNICAL_HINTS = (
    'வியல்', 'அறிவியல்', 'தொழில்', 'அரசியல்', 'பொருளாதார', 'வரலாறு',
    'இலக்கிய', 'இலக்கண', 'வேதியியல்', 'இயற்பியல்', 'உயிரியல்',
)


def _estimate_grade_for_wiki_word(
    stem: str,
    display_word: str,
    frequency: int,
    max_frequency: int,
    known_grade_map: Optional[Dict[str, int]] = None,
) -> Tuple[int, float, str]:
    known_grade_map = known_grade_map or {}
    if stem in known_grade_map:
        return max(1, min(12, int(known_grade_map[stem]))), 0.98, 'textbook_anchor'

    word = display_word or stem
    freq_ratio = (frequency / max_frequency) if max_frequency else 0.0

    if freq_ratio >= 0.020:
        grade, conf, reason = 3, 0.70, 'very_common_wikipedia'
    elif freq_ratio >= 0.006:
        grade, conf, reason = 5, 0.64, 'common_wikipedia'
    elif freq_ratio >= 0.0015:
        grade, conf, reason = 7, 0.58, 'moderate_wikipedia'
    elif frequency >= 25:
        grade, conf, reason = 9, 0.52, 'less_common_wikipedia'
    else:
        grade, conf, reason = 11, 0.45, 'rare_wikipedia'

    concept = classify_concept(stem)
    if concept in {'daily_life', 'nature_environment', 'health_body'}:
        grade -= 1
        conf += 0.04
        reason += '+daily_concept'
    elif concept in {'science_technology', 'society_civics', 'history_culture', 'math_logic'}:
        grade += 1
        conf += 0.04
        reason += '+academic_concept'

    if len(word) >= 12:
        grade += 1
        conf += 0.03
        reason += '+long_word'
    if len(word) >= 16:
        grade += 1
        reason += '+very_long_word'
    if word.endswith(FORMAL_GRADE_SUFFIXES) or any(h in word for h in TECHNICAL_HINTS):
        grade += 1
        conf += 0.04
        reason += '+formal_or_technical'

    if stem in STOPWORDS:
        grade = min(grade, 3)
        conf = max(conf, 0.70)
        reason = 'common_function_word'

    return max(1, min(12, grade)), round(min(conf, 0.92), 2), reason


def estimate_wiki_grade_levels(
    known_grade_map: Optional[Dict[str, int]] = None,
    batch_size: int = 5000,
) -> Dict:
    """Estimate and store Std 1-12 levels for every word in wiki_corpus.db.

    Grades are approximate. `grade_reason` and `confidence` distinguish
    textbook-anchored estimates from broader frequency/morphology heuristics.
    """
    init_wiki_db()
    known_grade_map = known_grade_map or {}
    conn = get_wiki_db()
    max_frequency = conn.execute(
        'SELECT COALESCE(MAX(frequency), 0) FROM wiki_words'
    ).fetchone()[0] or 0
    rows = conn.execute('''
        SELECT stem, display_word, frequency
        FROM wiki_words
        ORDER BY frequency DESC
    ''').fetchall()
    now = datetime.datetime.now().isoformat()
    by_grade: Counter[int] = Counter()
    by_reason: Counter[str] = Counter()
    updated = 0
    anchored = 0
    for r in rows:
        grade, confidence, reason = _estimate_grade_for_wiki_word(
            r['stem'], r['display_word'] or r['stem'], int(r['frequency'] or 0),
            int(max_frequency or 0), known_grade_map
        )
        if reason == 'textbook_anchor':
            anchored += 1
        by_grade[grade] += 1
        by_reason[reason.split('+')[0]] += 1
        conn.execute('''
            UPDATE wiki_words
            SET inferred_grade=?, confidence=?, grade_reason=?,
                grade_estimated_at=?, updated_at=?
            WHERE stem=?
        ''', (grade, confidence, reason, now, now, r['stem']))
        updated += 1
        if updated % int(batch_size) == 0:
            conn.commit()
    conn.execute('''
        INSERT INTO wiki_imports
          (dump_path, dump_type, article_count, stem_count, imported_at, status)
        VALUES (?,?,?,?,?,?)
    ''', (WIKI_DB, 'grade_estimation', 0, updated, now, 'active'))
    conn.commit()
    conn.close()
    return {
        'ok': True,
        'database': WIKI_DB,
        'updated': updated,
        'textbook_anchored': anchored,
        'heuristic_estimated': updated - anchored,
        'by_grade': [{'grade': g, 'cnt': c} for g, c in sorted(by_grade.items())],
        'by_reason': [{'reason': k, 'cnt': c} for k, c in by_reason.most_common()],
    }


def lookup_wiki_words(stems: List[str], limit: int = 500) -> Dict[str, Dict]:
    """Return estimated Wikipedia grade info for the requested stems."""
    init_wiki_db()
    stems = [s for s in dict.fromkeys(stems or []) if s][:int(limit)]
    if not stems:
        return {}
    conn = get_wiki_db()
    out: Dict[str, Dict] = {}
    for i in range(0, len(stems), 200):
        chunk = stems[i:i + 200]
        qs = ','.join(['?'] * len(chunk))
        rows = conn.execute(f'''
            SELECT stem, display_word, inferred_grade, confidence, grade_reason, frequency
            FROM wiki_words
            WHERE stem IN ({qs})
        ''', chunk).fetchall()
        for r in rows:
            out[r['stem']] = dict(r)
    conn.close()
    return out


def backfill_wiki_db_from_library(limit: int = 0) -> Dict:
    """Copy existing Wikipedia rows from word_library.db into data/wiki_corpus.db.

    This keeps older installs compatible after introducing the separate
    Wikipedia corpus database.
    """
    init_library_db()
    init_wiki_db()
    lib_conn = get_lib_db()
    wiki_conn = get_wiki_db()
    query = '''
        SELECT stem, display_word, grade_level, frequency, example
        FROM word_library
        WHERE grade_source='wikipedia'
        ORDER BY frequency DESC
    '''
    if limit:
        query += f' LIMIT {int(limit)}'
    rows = lib_conn.execute(query).fetchall()
    copied = 0
    for r in rows:
        upsert_wiki_word(
            wiki_conn,
            stem=r['stem'],
            display_word=r['display_word'] or r['stem'],
            inferred_grade=int(r['grade_level'] or 8),
            confidence=0.0,
            frequency=int(r['frequency'] or 1),
            example=r['example'],
            source_dump='backfilled_from_word_library',
        )
        copied += 1
        if copied % 5000 == 0:
            wiki_conn.commit()
    now = datetime.datetime.now().isoformat()
    wiki_conn.execute('''
        INSERT INTO wiki_imports
          (dump_path, dump_type, article_count, stem_count, imported_at, status)
        VALUES (?,?,?,?,?,?)
    ''', ('word_library.db', 'backfill', 0, copied, now, 'active'))
    wiki_conn.commit()
    lib_conn.close()
    wiki_conn.close()
    return {'copied': copied, 'database': WIKI_DB}


# ── Concept classification ────────────────────────────────────────────────────

def _tamil_prefix(stem: str, n: int) -> str:
    chars = [c for c in stem if '\u0B80' <= c <= '\u0BFF']
    return ''.join(chars[:n])


def classify_concept(stem: str) -> str:
    """Classify a Tamil stem into a concept category using seed matching."""
    if not stem or stem in STOPWORDS:
        return 'general'
    for category, seeds in CONCEPT_SEEDS.items():
        if stem in seeds:
            return category
        # 3-char Tamil prefix match
        sp = _tamil_prefix(stem, 3)
        if sp:
            for seed in seeds:
                if _tamil_prefix(seed, 3) == sp:
                    return category
    return 'general'


# ── Grade inference for non-textbook words ────────────────────────────────────

def infer_grade_from_context(
    stem: str,
    context_stems: List[str],
    known_grade_map: Dict[str, int],
    default: int = 8,
) -> Tuple[int, float]:
    """
    Infer grade for an unknown word by averaging the grades of words
    it co-occurs with in the same sentence.
    Returns (inferred_grade, confidence 0-1).
    """
    grades = [known_grade_map[s] for s in context_stems
              if s in known_grade_map and s != stem]
    if not grades:
        return default, 0.0
    avg = sum(grades) / len(grades)
    confidence = min(len(grades) / 5.0, 1.0)  # 5+ context words = full confidence
    return round(avg), confidence


# ── Core: add / update a word in the library ─────────────────────────────────

def upsert_word(
    conn: sqlite3.Connection,
    stem: str,
    display_word: str,
    grade_level: int,
    grade_source: str,
    frequency: int,
    source_name: str,
    example: Optional[str] = None,
    example_source: Optional[str] = None,
    part_of_speech: Optional[str] = None,
    definition: Optional[str] = None,
    confirmed: int = 0,
) -> None:
    """
    Insert or update a word in the library.
    Rules:
      - 'manual' source always wins (highest trust)
      - 'textbook' beats 'children_book' beats 'wikipedia'
      - Lower grade wins when same source type
      - Frequency is always accumulated
    """
    SOURCE_PRIORITY = {'manual': 0, 'textbook': 1, 'children_book': 2,
                       'wikipedia': 3, 'wiktionary': 4, 'inferred': 5}

    now = datetime.datetime.now().isoformat()
    concept = classify_concept(stem)

    existing = conn.execute(
        'SELECT * FROM word_library WHERE stem = ?', (stem,)
    ).fetchone()

    if existing is None:
        sources = [{'source': source_name, 'grade': grade_level, 'freq': frequency}]
        conn.execute('''
            INSERT INTO word_library
              (stem, display_word, grade_level, grade_source, frequency,
               concept, example, example_source, part_of_speech, definition,
               confirmed, sources_json, added_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (stem, display_word, grade_level, grade_source, frequency,
              concept, example, example_source, part_of_speech, definition,
              confirmed, json.dumps(sources), now, now))
        return

    # Merge with existing
    existing = dict(existing)
    try:
        sources = json.loads(existing.get('sources_json') or '[]')
    except Exception:
        sources = []

    # Update source list
    src_entry = next((s for s in sources if s.get('source') == source_name), None)
    if src_entry:
        src_entry['freq']  = src_entry.get('freq', 0) + frequency
        src_entry['grade'] = min(src_entry.get('grade', grade_level), grade_level)
    else:
        sources.append({'source': source_name, 'grade': grade_level, 'freq': frequency})

    new_freq = existing.get('frequency', 0) + frequency

    # Grade: lower source-priority wins; within same priority, lower grade wins
    existing_prio = SOURCE_PRIORITY.get(existing.get('grade_source', 'inferred'), 5)
    new_prio      = SOURCE_PRIORITY.get(grade_source, 5)

    if new_prio < existing_prio or (new_prio == existing_prio and
                                     grade_level < (existing.get('grade_level') or 99)):
        new_grade        = grade_level
        new_grade_source = grade_source
        new_confirmed    = confirmed
    else:
        new_grade        = existing.get('grade_level', grade_level)
        new_grade_source = existing.get('grade_source', grade_source)
        new_confirmed    = max(existing.get('confirmed', 0), confirmed)

    # Use new example if the existing one is None
    new_example = existing.get('example') or example
    new_ex_src  = existing.get('example_source') or example_source
    new_pos     = existing.get('part_of_speech') or part_of_speech
    new_def     = existing.get('definition') or definition

    conn.execute('''
        UPDATE word_library SET
          grade_level=?, grade_source=?, frequency=?, concept=?,
          example=?, example_source=?, part_of_speech=?,
          definition=?, confirmed=?, sources_json=?, updated_at=?
        WHERE stem=?
    ''', (new_grade, new_grade_source, new_freq, concept,
          new_example, new_ex_src, new_pos,
          new_def, new_confirmed, json.dumps(sources), now, stem))


# ── Source 1: Build from existing textbook DB ─────────────────────────────────

def build_from_textbooks(
    main_db_path: str,
    stem_fn,
    progress_cb=None,
) -> Dict:
    """
    Populate the word library from the already-loaded grade books DB.
    This is the fastest and most reliable source.
    """
    init_library_db()

    main_conn = sqlite3.connect(main_db_path)
    main_conn.row_factory = sqlite3.Row

    rows = main_conn.execute(
        'SELECT stem, first_grade FROM word_grade_map'
    ).fetchall()

    # Get display words (most common surface form per stem)
    gw_rows = main_conn.execute(
        'SELECT grade, word FROM grade_words'
    ).fetchall()

    # Get example sentences from grade_files text if available
    # (We store raw_text in pending_extractions but not permanently — skip for now)

    main_conn.close()

    if not rows:
        return {'added': 0, 'updated': 0, 'error': 'No textbook data loaded yet'}

    lib_conn = get_lib_db()
    added = updated = 0
    total = len(rows)

    for i, row in enumerate(rows):
        stem  = row['stem']
        grade = row['first_grade']
        if not stem or not grade:
            continue

        before = lib_conn.execute(
            'SELECT stem FROM word_library WHERE stem=?', (stem,)
        ).fetchone()

        upsert_word(
            lib_conn, stem=stem,
            display_word=stem,  # will be refined when we see surface forms
            grade_level=grade, grade_source='textbook',
            frequency=1, source_name='textbook_db',
        )

        if before:
            updated += 1
        else:
            added += 1

        if progress_cb and i % 500 == 0:
            progress_cb(i, total, 'textbooks')

    # Update display words from grade_words surface forms
    for r in gw_rows:
        stem = r['word']
        existing = lib_conn.execute(
            'SELECT display_word FROM word_library WHERE stem=?', (stem,)
        ).fetchone()
        if existing and existing['display_word'] == stem:
            # Already set, skip
            pass

    lib_conn.commit()
    lib_conn.close()

    return {'added': added, 'updated': updated, 'source': 'textbooks'}


# ── Source 2: Import a children's book / any PDF ─────────────────────────────

def import_from_book(
    filepath: str,
    source_name: str,
    grade_hint: Optional[int],
    extract_fn,
    stem_fn,
    known_grade_map: Dict[str, int],
    progress_cb=None,
) -> Dict:
    """
    Extract words from a PDF/TXT and add to the library.
    If grade_hint is given, use it. Otherwise infer from co-occurrence.
    Extracts example sentences for each new word.
    """
    init_library_db()

    text = extract_fn(filepath)
    if not text:
        return {'error': f'No text extracted from {os.path.basename(filepath)}'}

    # Split into sentences to get context + examples
    sentences = [s.strip() for s in SENT_RE.split(text) if s.strip()]
    if not sentences:
        sentences = [text.strip()]

    # Build stem frequency and context map
    stem_freq:    Counter  = Counter()
    stem_display: Dict[str, str] = {}
    stem_example: Dict[str, Tuple[str, str]] = {}  # stem -> (example, source)
    stem_context: Dict[str, List[List[str]]] = defaultdict(list)  # stem -> [[context_stems]]

    for sent in sentences:
        words = TAMIL_RE.findall(sent)
        if not words:
            continue
        stemmed = [(w, stem_fn(w)) for w in words]
        sent_stems = [s for _, s in stemmed if s]

        for w, s in stemmed:
            if not s or s in STOPWORDS or len(s) < 2:
                continue
            stem_freq[s] += 1
            if s not in stem_display:
                stem_display[s] = w
            if s not in stem_example and len(sent) < 200:
                stem_example[s] = (sent, source_name)
            stem_context[s].append([cs for cs in sent_stems if cs != s])

    lib_conn = get_lib_db()
    added = updated = 0
    total = len(stem_freq)

    for i, (stem, freq) in enumerate(stem_freq.most_common()):
        # Determine grade
        if grade_hint:
            grade = grade_hint
            confidence = 1.0
            gsource = 'children_book'
        else:
            # Infer from context
            all_context = [cs for ctx in stem_context[stem] for cs in ctx]
            grade, confidence = infer_grade_from_context(
                stem, all_context, known_grade_map
            )
            gsource = 'inferred'

        example, ex_src = stem_example.get(stem, (None, None))

        before = lib_conn.execute(
            'SELECT stem FROM word_library WHERE stem=?', (stem,)
        ).fetchone()

        upsert_word(
            lib_conn, stem=stem,
            display_word=stem_display.get(stem, stem),
            grade_level=grade, grade_source=gsource,
            frequency=freq, source_name=source_name,
            example=example, example_source=ex_src,
        )

        if before:
            updated += 1
        else:
            added += 1

        if progress_cb and i % 200 == 0:
            progress_cb(i, total, source_name)

    # Register source
    now = datetime.datetime.now().isoformat()
    lib_conn.execute('''
        INSERT OR REPLACE INTO library_sources
          (name, source_type, grade, file_path, word_count, added_at, status)
        VALUES (?,?,?,?,?,?,?)
    ''', (source_name, 'children_book' if not grade_hint else 'graded_book',
          grade_hint, filepath, len(stem_freq), now, 'active'))

    lib_conn.commit()
    lib_conn.close()

    return {
        'added': added, 'updated': updated,
        'total_stems': len(stem_freq),
        'source': source_name,
    }


# ── Source 3: Tamil Wikipedia (one-time download + parse) ────────────────────

WIKI_DUMP_URL  = 'https://dumps.wikimedia.org/tawiki/latest/tawiki-latest-pages-articles.xml.bz2'
WIKI_ABS_URL   = 'https://dumps.wikimedia.org/tawiki/latest/tawiki-latest-abstract.xml.gz'
WIKT_DUMP_URL  = 'https://dumps.wikimedia.org/tawiktionary/latest/tawiktionary-latest-pages-articles.xml.bz2'

WIKI_DOWNLOAD_DIR = Path('data/wiki_dumps')


def download_wiki_dump(dump_type: str = 'abstracts', progress_cb=None) -> Optional[str]:
    """
    Download Tamil Wikipedia or Wiktionary dump.
    dump_type: 'abstracts' (small, ~30 MB) | 'full' (~500 MB) | 'wiktionary'
    Returns local file path or None on failure.
    """
    WIKI_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    url_map = {
        'abstracts':  (WIKI_ABS_URL,   'tawiki_abstracts.xml.gz'),
        'full':       (WIKI_DUMP_URL,   'tawiki_full.xml.bz2'),
        'wiktionary': (WIKT_DUMP_URL,   'tawiktionary.xml.bz2'),
    }
    if dump_type not in url_map:
        return None

    url, filename = url_map[dump_type]
    dest = WIKI_DOWNLOAD_DIR / filename

    if dest.exists() and dest.stat().st_size > 1000:
        return str(dest)  # Already downloaded

    try:
        def _report(count, block_size, total):
            if progress_cb and total > 0:
                pct = count * block_size / total * 100
                progress_cb(min(pct, 100), f'Downloading {filename}')

        urllib.request.urlretrieve(url, dest, reporthook=_report)
        return str(dest)
    except Exception as e:
        if dest.exists():
            dest.unlink()
        return None


def import_from_wiki_dump(
    dump_path: str,
    stem_fn,
    known_grade_map: Dict[str, int],
    max_articles: int = 0,
    progress_cb=None,
) -> Dict:
    """
    Parse a downloaded Tamil Wikipedia dump and extract vocabulary.
    Grades are inferred from co-occurrence with known textbook words.
    max_articles=0 means process all.
    """
    init_library_db()
    init_wiki_db()

    import xml.etree.ElementTree as ET
    from io import BytesIO, TextIOWrapper

    dump_path = Path(dump_path)
    is_gz  = dump_path.suffix == '.gz'
    is_bz2 = dump_path.suffix == '.bz2'

    stem_data: Dict[str, Dict] = {}  # stem -> {freq, display, examples, contexts}
    article_count = 0

    def _process_article(title: str, text: str) -> None:
        nonlocal article_count
        # Skip non-content pages
        if ':' in title and not title.startswith('தமிழ்'):
            return

        # Clean wiki markup
        text = re.sub(r'\{\{[^}]*\}\}', '', text)
        text = re.sub(r'\[\[(?:[^|\]]*\|)?([^\]]*)\]\]', r'\1', text)
        text = re.sub(r'[=\|\*#<>]', ' ', text)

        sentences = [s.strip() for s in SENT_RE.split(text) if s.strip()]
        for sent in sentences[:20]:  # Limit per article
            words = TAMIL_RE.findall(sent)
            if len(words) < 3:
                continue
            stemmed = [(w, stem_fn(w)) for w in words]
            sent_stems = [s for _, s in stemmed if s and len(s) >= 2]

            for w, s in stemmed:
                if not s or s in STOPWORDS or len(s) < 2:
                    continue
                if s not in stem_data:
                    stem_data[s] = {
                        'display': w, 'freq': 0,
                        'examples': [], 'contexts': [],
                    }
                sd = stem_data[s]
                sd['freq'] += 1
                if len(sd['examples']) < 2 and len(sent) < 150:
                    sd['examples'].append(sent)
                sd['contexts'].extend([cs for cs in sent_stems if cs != s])

        article_count += 1
        if max_articles and article_count >= max_articles:
            raise StopIteration

    try:
        if is_gz:
            opener = lambda: gzip.open(dump_path, 'rb')
        elif is_bz2:
            opener = lambda: bz2.open(dump_path, 'rb')
        else:
            opener = lambda: open(dump_path, 'rb')

        with opener() as f:
            current_title = ''
            current_text  = []
            in_text = False

            for line in f:
                try:
                    line = line.decode('utf-8', errors='ignore')
                except Exception:
                    continue

                if '<title>' in line:
                    m = re.search(r'<title>(.*?)</title>', line)
                    if m:
                        current_title = html.unescape(m.group(1))
                        current_text  = []
                elif '<text' in line:
                    in_text = True
                    m = re.search(r'<text[^>]*>(.*)', line)
                    if m:
                        current_text.append(html.unescape(m.group(1)))
                elif '</text>' in line:
                    if in_text:
                        m = re.search(r'(.*?)</text>', line)
                        if m:
                            current_text.append(html.unescape(m.group(1)))
                    in_text = False
                    try:
                        _process_article(current_title, '\n'.join(current_text))
                    except StopIteration:
                        break
                    if progress_cb and article_count % 1000 == 0:
                        progress_cb(article_count, max_articles or article_count,
                                    'wikipedia')
                elif in_text:
                    current_text.append(html.unescape(line))

    except StopIteration:
        pass
    except Exception as e:
        return {'error': str(e)}

    # Write to both databases:
    # 1) word_library.db for grade-browsing/search UI
    # 2) data/wiki_corpus.db as the separate Wikipedia corpus store
    lib_conn = get_lib_db()
    wiki_conn = get_wiki_db()
    added = updated = 0

    for stem, sd in stem_data.items():
        contexts = sd['contexts'][:50]
        grade, conf = infer_grade_from_context(stem, contexts, known_grade_map)

        before = lib_conn.execute(
            'SELECT stem FROM word_library WHERE stem=?', (stem,)
        ).fetchone()

        example = sd['examples'][0] if sd['examples'] else None
        upsert_word(
            lib_conn, stem=stem,
            display_word=sd['display'],
            grade_level=grade, grade_source='wikipedia',
            frequency=sd['freq'], source_name='tamil_wikipedia',
            example=example, example_source='Tamil Wikipedia',
        )
        upsert_wiki_word(
            wiki_conn,
            stem=stem,
            display_word=sd['display'],
            inferred_grade=grade,
            confidence=conf,
            frequency=sd['freq'],
            example=example,
            source_dump=str(dump_path),
        )
        if before: updated += 1
        else: added += 1

    now = datetime.datetime.now().isoformat()
    lib_conn.execute('''
        INSERT OR REPLACE INTO library_sources
          (name, source_type, grade, file_path, word_count, added_at, status)
        VALUES (?,?,?,?,?,?,?)
    ''', ('tamil_wikipedia', 'wikipedia', None, str(dump_path),
          len(stem_data), now, 'active'))
    wiki_conn.execute('''
        INSERT INTO wiki_imports
          (dump_path, dump_type, article_count, stem_count, imported_at, status)
        VALUES (?,?,?,?,?,?)
    ''', (str(dump_path), dump_path.suffix.lstrip('.') or 'wiki',
          article_count, len(stem_data), now, 'active'))

    lib_conn.commit()
    lib_conn.close()
    wiki_conn.commit()
    wiki_conn.close()

    return {'added': added, 'updated': updated,
            'articles': article_count, 'stems': len(stem_data),
            'wiki_database': WIKI_DB}


# ── Manual entry ──────────────────────────────────────────────────────────────

def manual_entry(
    stem: str,
    display_word: str,
    grade_level: int,
    definition: Optional[str] = None,
    part_of_speech: Optional[str] = None,
    example: Optional[str] = None,
) -> Dict:
    """Add or override a word manually (teacher/editor entry)."""
    init_library_db()
    lib_conn = get_lib_db()
    upsert_word(
        lib_conn, stem=stem, display_word=display_word or stem,
        grade_level=grade_level, grade_source='manual',
        frequency=1, source_name='manual',
        definition=definition, part_of_speech=part_of_speech,
        example=example, confirmed=1,
    )
    lib_conn.commit()
    lib_conn.close()
    return {'ok': True, 'stem': stem, 'grade': grade_level}


# ── Query API ─────────────────────────────────────────────────────────────────

def search_word(query: str, limit: int = 20) -> List[Dict]:
    """Search the library by stem or display word prefix."""
    init_library_db()
    lib_conn = get_lib_db()
    rows = lib_conn.execute('''
        SELECT * FROM word_library
        WHERE stem LIKE ? OR display_word LIKE ?
        ORDER BY confirmed DESC, frequency DESC
        LIMIT ?
    ''', (f'{query}%', f'{query}%', limit)).fetchall()
    lib_conn.close()
    return [dict(r) for r in rows]


def get_by_grade(grade: int, concept: Optional[str] = None,
                 confirmed_only: bool = False,
                 limit: int = 500, offset: int = 0) -> List[Dict]:
    """Get all words for a specific grade, optionally filtered by concept."""
    init_library_db()
    lib_conn = get_lib_db()
    filters = ['grade_level = ?']
    params: List = [grade]
    if concept and concept != 'all':
        filters.append('concept = ?')
        params.append(concept)
    if confirmed_only:
        filters.append('confirmed = 1')
    where = ' AND '.join(filters)
    rows = lib_conn.execute(
        f'SELECT * FROM word_library WHERE {where} '
        f'ORDER BY confirmed DESC, frequency DESC LIMIT ? OFFSET ?',
        params + [limit, offset]
    ).fetchall()
    total = lib_conn.execute(
        f'SELECT COUNT(*) FROM word_library WHERE {where}', params
    ).fetchone()[0]
    lib_conn.close()
    return {'words': [dict(r) for r in rows], 'total': total}


def get_stats() -> Dict:
    """Get library statistics."""
    init_library_db()
    lib_conn = get_lib_db()
    total   = lib_conn.execute('SELECT COUNT(*) FROM word_library').fetchone()[0]
    by_grade = lib_conn.execute(
        'SELECT grade_level, COUNT(*) as cnt FROM word_library '
        'GROUP BY grade_level ORDER BY grade_level'
    ).fetchall()
    by_concept = lib_conn.execute(
        'SELECT concept, COUNT(*) as cnt FROM word_library '
        'GROUP BY concept ORDER BY cnt DESC'
    ).fetchall()
    by_source = lib_conn.execute(
        'SELECT grade_source, COUNT(*) as cnt FROM word_library '
        'GROUP BY grade_source ORDER BY cnt DESC'
    ).fetchall()
    confirmed = lib_conn.execute(
        'SELECT COUNT(*) FROM word_library WHERE confirmed=1'
    ).fetchone()[0]
    sources = lib_conn.execute(
        'SELECT * FROM library_sources ORDER BY added_at DESC'
    ).fetchall()
    lib_conn.close()
    return {
        'total':      total,
        'confirmed':  confirmed,
        'by_grade':   [dict(r) for r in by_grade],
        'by_concept': [dict(r) for r in by_concept],
        'by_source':  [dict(r) for r in by_source],
        'sources':    [dict(r) for r in sources],
    }


def export_to_excel(output_path: str) -> str:
    """Export the full word library to Excel with one sheet per grade."""
    init_library_db()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    lib_conn = get_lib_db()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_sum = wb.create_sheet('Summary')
    stats = get_stats()
    ws_sum.append(['Tamil Word Library — Summary'])
    ws_sum.append([f'Total words: {stats["total"]}',
                   f'Teacher-confirmed: {stats["confirmed"]}'])
    ws_sum.append([])
    ws_sum.append(['Grade', 'Word Count'])
    for r in stats['by_grade']:
        ws_sum.append([GRADE_LABELS.get(r['grade_level'], f'Std {r["grade_level"]}'),
                       r['cnt']])
    ws_sum.append([])
    ws_sum.append(['Concept', 'Word Count'])
    for r in stats['by_concept']:
        ws_sum.append([r['concept'], r['cnt']])

    # One sheet per grade
    for grade in range(1, 13):
        result = get_by_grade(grade, limit=5000)
        words  = result['words']
        if not words:
            continue
        ws = wb.create_sheet(f'Std {grade}')
        header = ['Word', 'Stem', 'Grade', 'Concept', 'Part of Speech',
                  'Definition', 'Example Sentence', 'Source', 'Frequency', 'Confirmed']
        ws.append(header)
        # Style header
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = PatternFill('solid', fgColor='1D9E75')
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        for w in words:
            ws.append([
                w.get('display_word') or w['stem'],
                w['stem'],
                GRADE_LABELS.get(w['grade_level'], str(w['grade_level'])),
                w.get('concept', ''),
                w.get('part_of_speech', ''),
                w.get('definition', ''),
                w.get('example', ''),
                w.get('grade_source', ''),
                w.get('frequency', 0),
                'Yes' if w.get('confirmed') else 'No',
            ])
        # Column widths
        for col, width in enumerate([18,18,12,18,14,30,50,16,10,10], 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

    lib_conn.close()
    wb.save(output_path)
    return output_path
